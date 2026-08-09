"""Sync the Trading 212 portfolio into portfolio.xlsx.

Run from the project root:

    .venv/bin/python -m t212.sync

Writes:
    portfolio.xlsx      Summary / Holdings / History / HoldingsHistory sheets
    data/portfolio.json Latest snapshot, machine-readable (for a dashboard)

History sheets are append-only: one row per sync per day. Re-running on
the same day overwrites that day's row rather than adding a duplicate.
"""

from __future__ import annotations

import argparse
import json
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

import pandas as pd

from t212.client import T212Error, Trading212

ROOT = Path(__file__).resolve().parent.parent
WORKBOOK = ROOT / "portfolio.xlsx"
JSON_SNAPSHOT = ROOT / "data" / "portfolio.json"
DIVIDENDS_JSON = ROOT / "data" / "dividends.json"

MONEY = "#,##0.00"
PERCENT = "0.00%"
SHARES = "#,##0.####"

# Sheet names. Note: Excel reserves "History" for shared-workbook change
# tracking and refuses to open a file that uses it, hence "Snapshots".
SUMMARY_SHEET = "Summary"
HOLDINGS_SHEET = "Holdings"
HISTORY_SHEET = "Snapshots"
HOLDINGS_HISTORY_SHEET = "HoldingsSnapshots"

# Older workbooks used these names; read them so history survives a rename.
LEGACY_SHEETS = {
    HISTORY_SHEET: "History",
    HOLDINGS_HISTORY_SHEET: "HoldingsHistory",
}

ALL_SHEETS = [SUMMARY_SHEET, HOLDINGS_SHEET, HISTORY_SHEET, HOLDINGS_HISTORY_SHEET]


# ---------------------------------------------------------------- shaping


def build_summary(summary: dict) -> dict:
    """Flatten the nested account summary into one row of metrics."""
    cash = summary.get("cash") or {}
    inv = summary.get("investments") or {}

    total_cost = float(inv.get("totalCost") or 0.0)
    unrealised = float(inv.get("unrealizedProfitLoss") or 0.0)

    return {
        "Account ID": summary.get("id"),
        "Currency": summary.get("currency"),
        "Total account value": float(summary.get("totalValue") or 0.0),
        "Investments value": float(inv.get("currentValue") or 0.0),
        "Invested (cost basis)": total_cost,
        "Unrealised P/L": unrealised,
        "Unrealised P/L %": (unrealised / total_cost) if total_cost else 0.0,
        "Realised P/L (all time)": float(inv.get("realizedProfitLoss") or 0.0),
        "Cash available to trade": float(cash.get("availableToTrade") or 0.0),
        "Cash in pies": float(cash.get("inPies") or 0.0),
        "Cash reserved for orders": float(cash.get("reservedForOrders") or 0.0),
    }


def build_holdings(positions: list[dict]) -> pd.DataFrame:
    """One row per open position, sorted by current value descending."""
    rows = []
    for position in positions:
        instrument = position.get("instrument") or {}
        wallet = position.get("walletImpact") or {}

        cost = float(wallet.get("totalCost") or 0.0)
        value = float(wallet.get("currentValue") or 0.0)
        pnl = float(wallet.get("unrealizedProfitLoss") or (value - cost))

        rows.append(
            {
                "Ticker": instrument.get("ticker"),
                "Name": instrument.get("name"),
                "ISIN": instrument.get("isin"),
                "Instrument currency": instrument.get("currency"),
                "Quantity": float(position.get("quantity") or 0.0),
                "Avg price paid": float(position.get("averagePricePaid") or 0.0),
                "Current price": float(position.get("currentPrice") or 0.0),
                "Total cost": cost,
                "Current value": value,
                "Unrealised P/L": pnl,
                "Unrealised P/L %": (pnl / cost) if cost else 0.0,
                "FX impact": float(wallet.get("fxImpact") or 0.0),
                "Qty in pies": float(position.get("quantityInPies") or 0.0),
                "Opened": _as_date(position.get("createdAt")),
            }
        )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return frame

    total_value = frame["Current value"].sum()
    frame["Weight %"] = frame["Current value"] / total_value if total_value else 0.0
    frame = frame.sort_values("Current value", ascending=False).reset_index(drop=True)

    # Put weight next to value rather than at the far right.
    columns = list(frame.columns)
    columns.insert(columns.index("Current value") + 1, columns.pop(columns.index("Weight %")))
    return frame[columns]


def build_dividends(raw: list[dict]) -> pd.DataFrame:
    """Normalise the dividends payload into one row per payment.

    The API has used several key spellings across versions, so each field
    is resolved from a list of candidates rather than a fixed name.
    """
    def pick(row: dict, *names, default=None):
        for name in names:
            if name in row and row[name] not in (None, ""):
                return row[name]
        return default

    rows = []
    for item in raw or []:
        amount = pick(item, "amount", "amountInEuro", "grossAmountPerShare", default=0)
        rows.append(
            {
                "Ticker": pick(item, "ticker", "instrumentTicker", default=""),
                "Paid on": _as_date(
                    pick(item, "paidOn", "paidOnDate", "date", "time", default="")
                ),
                "Amount": float(amount or 0.0),
                "Quantity": float(pick(item, "quantity", "shares", default=0) or 0.0),
                "Per share": float(
                    pick(item, "grossAmountPerShare", "amountPerShare", default=0) or 0.0
                ),
                "Type": pick(item, "type", default=""),
                "Reference": pick(item, "reference", "id", default=""),
            }
        )

    frame = pd.DataFrame(rows)
    if frame.empty:
        return pd.DataFrame(
            columns=["Ticker", "Paid on", "Amount", "Quantity", "Per share",
                     "Type", "Reference"]
        )
    return frame.sort_values("Paid on", ascending=False).reset_index(drop=True)


def _as_date(value) -> str:
    if not value:
        return ""
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime(
            "%Y-%m-%d"
        )
    except ValueError:
        return str(value)


# ------------------------------------------------------------- persistence


def _read_sheet(path: Path, sheet: str) -> pd.DataFrame:
    """Read a sheet, falling back to its pre-rename name if needed."""
    if not path.exists():
        return pd.DataFrame()
    for name in (sheet, LEGACY_SHEETS.get(sheet)):
        if not name:
            continue
        try:
            return pd.read_excel(path, sheet_name=name)
        except (ValueError, KeyError):  # sheet absent
            continue
    return pd.DataFrame()


def _append_snapshot(existing: pd.DataFrame, new: pd.DataFrame) -> pd.DataFrame:
    """Append new rows, replacing any rows already logged for that date."""
    if existing.empty:
        return new
    existing = existing.copy()
    existing["Date"] = existing["Date"].astype(str).str.slice(0, 10)
    today = new["Date"].iloc[0]
    existing = existing[existing["Date"] != today]
    return pd.concat([existing, new], ignore_index=True)


def write_workbook(
    summary_row: dict,
    holdings: pd.DataFrame,
    history: pd.DataFrame,
    holdings_history: pd.DataFrame,
    path: Path,
) -> None:
    summary_frame = pd.DataFrame(
        {"Metric": list(summary_row), "Value": list(summary_row.values())}
    )
    summary_frame.loc[len(summary_frame)] = [
        "Last updated",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    ]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary_frame.to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
        holdings.to_excel(writer, sheet_name=HOLDINGS_SHEET, index=False)
        history.to_excel(writer, sheet_name=HISTORY_SHEET, index=False)
        holdings_history.to_excel(
            writer, sheet_name=HOLDINGS_HISTORY_SHEET, index=False
        )
        _style(writer, holdings, history, holdings_history)


def _style(writer, holdings, history, holdings_history) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F3864")

    money_columns = {
        "Total cost", "Current value", "Unrealised P/L", "FX impact",
        "Avg price paid", "Current price", "Value", "Total value",
        "Investments value", "Invested (cost basis)", "Realised P/L",
        "Cash", "Cost",
    }
    percent_columns = {"Unrealised P/L %", "Weight %"}
    share_columns = {"Quantity", "Qty in pies"}

    frames = {
        HOLDINGS_SHEET: holdings,
        HISTORY_SHEET: history,
        HOLDINGS_HISTORY_SHEET: holdings_history,
    }

    for name in ALL_SHEETS:
        sheet = writer.book[name]
        sheet.freeze_panes = "A2"

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left")

        frame = frames.get(name)
        if frame is None or frame.empty:
            _autosize(sheet, get_column_letter)
            continue

        for index, column in enumerate(frame.columns, start=1):
            letter = get_column_letter(index)
            if column in percent_columns:
                fmt = PERCENT
            elif column in share_columns:
                fmt = SHARES
            elif column in money_columns:
                fmt = MONEY
            else:
                continue
            for cell in sheet[letter][1:]:
                cell.number_format = fmt

        _autosize(sheet, get_column_letter)

    # Summary values column: money format on the numeric rows.
    summary_sheet = writer.book[SUMMARY_SHEET]
    for row in summary_sheet.iter_rows(min_row=2, min_col=1, max_col=2):
        label, value = row[0].value, row[1]
        if isinstance(label, str) and label.endswith("%"):
            value.number_format = PERCENT
        elif isinstance(value.value, (int, float)) and label != "Account ID":
            value.number_format = MONEY


def _autosize(sheet, get_column_letter) -> None:
    for column_cells in sheet.columns:
        longest = max(
            (len(str(cell.value)) for cell in column_cells if cell.value is not None),
            default=0,
        )
        letter = get_column_letter(column_cells[0].column)
        sheet.column_dimensions[letter].width = min(max(longest + 2, 10), 42)


# -------------------------------------------------------------------- main


def sync(env_file: Path | None = None, workbook: Path = WORKBOOK) -> dict:
    client = Trading212.from_env(env_file, root=ROOT)

    summary = client.account_summary()
    time.sleep(1)  # positions allows 1 req/s
    positions = client.positions()

    # Dividends are paginated at 6 req/min, so this is the slow part.
    # A failure here must not lose the position sync we already have.
    try:
        time.sleep(1)
        dividends = build_dividends(client.dividends())
    except Exception as error:  # noqa: BLE001
        print(f"  dividends fetch failed ({str(error)[:120]}) — keeping previous file")
        dividends = pd.DataFrame()

    summary_row = build_summary(summary)
    holdings = build_holdings(positions)
    today = datetime.now().strftime("%Y-%m-%d")

    # Map tickers to display names so the dividends page reads properly.
    if not dividends.empty and not holdings.empty:
        names = dict(zip(holdings["Ticker"], holdings["Name"]))
        dividends["Name"] = dividends["Ticker"].map(names).fillna(dividends["Ticker"])
        currency = summary_row.get("Currency") or "GBP"
        dividends["Currency"] = currency

    if not dividends.empty:
        DIVIDENDS_JSON.parent.mkdir(parents=True, exist_ok=True)
        DIVIDENDS_JSON.write_text(
            json.dumps(
                {
                    "generatedAt": datetime.now(timezone.utc).isoformat(),
                    "currency": summary_row.get("Currency") or "GBP",
                    "dividends": dividends.to_dict(orient="records"),
                },
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )

    history_row = pd.DataFrame(
        [
            {
                "Date": today,
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Total value": summary_row["Total account value"],
                "Investments value": summary_row["Investments value"],
                "Invested (cost basis)": summary_row["Invested (cost basis)"],
                "Unrealised P/L": summary_row["Unrealised P/L"],
                "Unrealised P/L %": summary_row["Unrealised P/L %"],
                "Realised P/L": summary_row["Realised P/L (all time)"],
                "Cash": summary_row["Cash available to trade"]
                + summary_row["Cash in pies"]
                + summary_row["Cash reserved for orders"],
                "Positions": len(holdings),
            }
        ]
    )

    if holdings.empty:
        holdings_rows = pd.DataFrame(
            columns=["Date", "Ticker", "Quantity", "Price", "Value", "Cost", "Unrealised P/L"]
        )
    else:
        holdings_rows = pd.DataFrame(
            {
                "Date": today,
                "Ticker": holdings["Ticker"],
                "Quantity": holdings["Quantity"],
                "Price": holdings["Current price"],
                "Value": holdings["Current value"],
                "Cost": holdings["Total cost"],
                "Unrealised P/L": holdings["Unrealised P/L"],
            }
        )

    history = _append_snapshot(_read_sheet(workbook, "History"), history_row)
    holdings_history = _append_snapshot(
        _read_sheet(workbook, "HoldingsHistory"), holdings_rows
    )

    write_workbook(summary_row, holdings, history, holdings_history, workbook)

    JSON_SNAPSHOT.parent.mkdir(parents=True, exist_ok=True)
    JSON_SNAPSHOT.write_text(
        json.dumps(
            {
                "generatedAt": datetime.now(timezone.utc).isoformat(),
                "summary": summary_row,
                "holdings": holdings.to_dict(orient="records"),
                "history": history.to_dict(orient="records"),
            },
            indent=2,
            default=str,
        ),
        encoding="utf-8",
    )

    return summary_row


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Trading 212 into Excel.")
    parser.add_argument("--env-file", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=WORKBOOK)
    args = parser.parse_args()

    try:
        summary_row = sync(args.env_file, args.output)
    except T212Error as error:
        print(f"Trading 212 sync failed: {error}", file=sys.stderr)
        return 1

    currency = summary_row["Currency"]
    print(
        f"Synced {args.output.name} — {currency} "
        f"{summary_row['Total account value']:,.2f} total, "
        f"{summary_row['Unrealised P/L']:+,.2f} unrealised "
        f"({summary_row['Unrealised P/L %']:+.2%})"
    )
    if DIVIDENDS_JSON.exists():
        payload = json.loads(DIVIDENDS_JSON.read_text(encoding="utf-8"))
        records = payload.get("dividends", [])
        total = sum(float(r.get("Amount") or 0) for r in records)
        print(f"Dividends: {len(records)} payments, {currency} {total:,.2f} total")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
